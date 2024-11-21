import pytest
from excsv.utils.excel import csv_to_workbook

import csv
from io import StringIO, BytesIO
from openpyxl import Workbook, load_workbook


@pytest.fixture
def input_file(tmp_path):
    d = tmp_path / "sub"
    d.mkdir()
    p = d / "input.csv"
    p.write_text(
        """
name,age
Alice,42
Bob,9
Chaz,101
""".strip()
    )
    return p


@pytest.fixture
def input_csv_text(input_file):
    return open(input_file, "r")




@pytest.mark.alpha
def test_csv_to_workbook(input_csv_text):
    csv_reader = csv.reader(input_csv_text)
    excel_bytes_io = csv_to_workbook(csv_reader, frozen_row=1, frozen_col=2)
    wb = load_workbook(excel_bytes_io)
    sheet = wb.active

    assert sheet["A1"].value == "name", "Verify the content matches the CSV input"
    assert sheet["B1"].value == "age"
    assert sheet["A2"].value == "Alice"
    assert sheet["B2"].value == "42"
    assert sheet["A3"].value == "Bob"
    assert sheet["B3"].value == "9"
    assert sheet["A4"].value == "Chaz"
    assert sheet["B4"].value == "101"

    assert sheet.freeze_panes == "B2", "Verify the default row/col is frozen"


@pytest.mark.alpha(
    """
    Write-only openpyxl does not appear to support sheet.dimensions attribute; i.e. have to write data, then apply filter
    maybe make write-only an 'efficient' option?

    For now, we set openpyxl.workbook(write_only) to False, and apply autofilter after the data has been written
    to the sheet
    """
)
def test_csv_to_workbook_auto_filter(input_csv_text):
    csv_reader = csv.reader(input_csv_text)
    excel_bytes_io = csv_to_workbook(csv_reader, frozen_row=1, frozen_col=2)
    wb = load_workbook(excel_bytes_io)
    sheet = wb.active

    assert sheet.auto_filter.ref == sheet.dimensions, "Verify auto filter is applied"


@pytest.mark.skip
def test_csv_to_workbook_set_default_header_styles(input_csv_text):
    assert 1 is 0

@pytest.mark.skip
def test_csv_to_workbook_enforce_max_column_width(input_csv_text):
    assert 1 is 0

@pytest.mark.skip
def test_csv_to_workbook_enforce_min_column_width(input_csv_text):
    assert 1 is 0
