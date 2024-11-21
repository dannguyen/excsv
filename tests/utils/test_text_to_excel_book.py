import pytest
from excsv.utils.excel import text_to_excel_book

import csv
from io import StringIO, BytesIO
from openpyxl import Workbook, load_workbook


@pytest.fixture
def input_file(tmp_path):
    d = tmp_path / "sub"
    d.mkdir()
    p = d / "input.csv"
    p.write_text(
        """
name,age
Alice,42
Bob,9
Chaz,101
""".strip()
    )
    return p


@pytest.fixture
def input_csv_text(input_file):
    return open(input_file, "r")


@pytest.mark.alpha
def test_text_to_excel_book(input_csv_text):
    csv_reader = csv.reader(input_csv_text)
    excel_bytes_io = text_to_excel_book(csv_reader, frozen_row=1, frozen_col=2)
    wb = load_workbook(excel_bytes_io)
    sheet = wb.active

    assert sheet["A1"].value == "name", "Verify the content matches the CSV input"
    assert sheet["B1"].value == "age"
    assert sheet["A2"].value == "Alice"
    assert sheet["B2"].value == "42"
    assert sheet["A3"].value == "Bob"
    assert sheet["B3"].value == "9"
    assert sheet["A4"].value == "Chaz"
    assert sheet["B4"].value == "101"

    assert sheet.freeze_panes == "B2", "Verify the default row/col is frozen"
    assert sheet.auto_filter.ref == sheet.dimensions, "Verify auto filter is applied"


@pytest.mark.alpha
def test_gemini_text_to_excel_book(input_file):
    """gemini's version of a pytest (blah!)"""

    # Simulate reading CSV data (replace with your actual CSV reading logic)
    with open(input_file, "r") as csv_file:
        csv_data = [row for row in csv.reader(csv_file)]

    # Call the text_to_excel_book function
    output_bytes = text_to_excel_book(csv_data)

    # Load the generated Excel workbook from the BytesIO object
    output_wb = load_workbook(output_bytes, read_only=True)
    output_sheet = output_wb.active

    # Assert expected behavior (adjust assertions based on your function's logic)
    assert output_sheet.cell(row=1, column=1).value == "name"
    assert output_sheet.cell(row=2, column=1).value == "Alice"
    # Add further assertions to check sheet dimensions, formatting, etc.
