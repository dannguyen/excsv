import pytest
from click.testing import CliRunner
from excsv.cli import cli


import csv
from io import StringIO, BytesIO
from openpyxl import Workbook, load_workbook


@pytest.fixture
def input_file(tmp_path):
    d = tmp_path / "sub"
    d.mkdir()
    p = d / "mydata.csv"
    p.write_text(
        """
name,age
Alice,42
Bob,9
Chaz,101
""".strip()
    )
    return p


def test_help_option():
    runner = CliRunner()
    result = runner.invoke(cli, ["--help"])
    assert result.exit_code == 0
    assert "Usage:" in result.output
    assert "--help " in result.output


def test_cli_version():
    runner = CliRunner()
    result = runner.invoke(cli, ["--version"])
    assert result.exit_code == 0
    assert "version" in result.output


def test_default_excsv_call_with_no_subcommand(input_file):
    """
    As shown in README.md
    $ excsv mydata.csv -o mysheet.xlsx
    """

    runner = CliRunner()

    with runner.isolated_filesystem():
        test_out_path = "mysheet.xlsx"

        result = runner.invoke(
            cli,
            [
                str(input_file),
                "--output-path",
                test_out_path,
            ],
        )
        assert result.exit_code == 0

        with open(test_out_path, "rb") as f:
            wb = load_workbook(test_out_path)
            sheet = wb.active

            assert (
                sheet["A1"].value == "name"
            ), "Verify the content matches the CSV input"
            assert sheet["B1"].value == "age"
            assert sheet["A2"].value == "Alice"
            assert sheet["B2"].value == "42"
            assert sheet["A3"].value == "Bob"
            assert sheet["B3"].value == "9"
            assert sheet["A4"].value == "Chaz"
            assert sheet["B4"].value == "101"


def test_default_excsv_call_stdout_with_no_subcommand(input_file):
    """
    As shown in README.md
    $ excsv mydata.csv > mysheet.xlsx
    """
    runner = CliRunner()

    with runner.isolated_filesystem():
        test_out_path = "mysheet.xlsx"

        with open(test_out_path, "wb") as output_file:
            result = runner.invoke(cli, [str(input_file)])
            output_file.write(result.stdout_bytes)

        assert result.exit_code == 0

        wb = load_workbook(test_out_path, read_only=True)
        sheet = wb.active

        assert sheet["A1"].value == "name"
        assert sheet["B1"].value == "age"
        assert sheet["A2"].value == "Alice"
        assert sheet["B2"].value == "42"


def test_default_excsv_call_stdin_stdout(input_file):
    """
    $ cat mydata.csv | excsv > mysheet.xlsx
    """
    runner = CliRunner()

    with runner.isolated_filesystem():
        test_out_path = "mysheet.xlsx"

        with input_file.open("r") as infile:
            result = runner.invoke(cli, input=infile.read())

        with open(test_out_path, "wb") as output_file:
            output_file.write(result.stdout_bytes)

        assert result.exit_code == 0

        wb = load_workbook(test_out_path, read_only=True)
        sheet = wb.active

        assert sheet["A1"].value == "name"
        assert sheet["B1"].value == "age"
        assert sheet["A2"].value == "Alice"
        assert sheet["B2"].value == "42"
