import csv
from io import BytesIO
import openpyxl
from typing import BinaryIO, TextIO


def text_to_excel_book(
    input_csv: csv.reader,
    frozen_row: int = 1,
    frozen_col: int = 2,
) -> BinaryIO:
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook(write_only=True)
    sheet = wb.create_sheet()

    # Simulate reading from a stream (e.g., a file or network stream)
    # Here, csv_stream is expected to be an iterable of CSV lines


    frozen_row_actual_num = frozen_row + 1
    frozen_col_letter = openpyxl.utils.cell.get_column_letter(frozen_col)
    sheet.freeze_panes = f"{frozen_col_letter}{frozen_row_actual_num}"


#    sheet.auto_filter.ref = sheet.dimensions

    for row in input_csv:
        # Append each row from the CSV to the Excel sheet
        sheet.append(row)



    excel_bytes = BytesIO()
    wb.save(excel_bytes)

    # Seek to the start of the BytesIO object so it can be read from the beginning
    excel_bytes.seek(0)

    # Return the BytesIO object
    return excel_bytes
